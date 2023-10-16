import argparse
import random
from pathlib import Path
from typing import Generator

import unicodedata
from openpyxl import load_workbook
import logging
from logging.handlers import RotatingFileHandler
from logging import StreamHandler


def create_class_user_files(path:str) -> None:
    """
    creates a user class bash file, a delteion file and a userlist
    :param path: The excel file
    """
    try:
        with open("create_users.sh", 'w') as file, open("delete_users.sh", 'w') as file_delete, \
                open("userlist.txt", 'w') as user_list_file:
            file.write("#! /bin/sh\n")
            file_delete.write("#! /bin/sh\n")
            file.write("groupadd klasse\n")
            file.write("groupadd lehrer\n")
            file.write("groupadd seminar\n")
            file.write("useradd -d /home/lehrer -g lehrer -c \"lehrer\" -s /bin/bash -G cdrom,plugdev,sambashare lehrer\n")
            file.write("useradd -d /home/seminar -g seminar -c \"seminar\" -s /bin/bash -G cdrom,plugdev,sambashare seminar\n")
            user_dict = dict()
            for class_name, room, kv in get_classes_from_class_file(path):
                class_name = str(class_name)
                room = str(room)
                kv = str(kv)
                username = replace_umlaute_and_remove_accents(f'k{str(class_name).lower()}')
                if username in user_dict:
                    logger.error(f"User duplicate: {username}. Programm is terminating!")
                    break
                command = f"useradd -d /home/klassen/{username} -g klasse -c \"{class_name} - {kv}\" -s /bin/bash -G cdrom,plugdev,sambashare {username}\n"
                password = f'{class_name}{get_random_pw_char()}{room[0:2]}{get_random_pw_char()}{kv}{get_random_pw_char()}'
                passwd_command = f'echo {username}:{password} | chpasswd\n'
                file.write(command)
                logger.debug(f"User created: {username}")
                file.write(passwd_command)
                logger.debug(f"User password created for {username}")
                file_delete.write(f'userdel -r {username}\n')
                logger.debug(f"User deletion written for {username}")
                user_list_file.write(f'username: {username}, password: {password}\n')
                logger.debug(f"User \"{username}\" added to list")
    except Exception:
        logger.critical("Couldn't write in one of the Files")

def get_random_pw_char():
    """
    :return: a random character from !%&(),._-=^#
    """
    return "!%&(),._-=^#"[random.randint(0, 11)]


def replace_umlaute_and_remove_accents(username):
    """
    Replaces accents and replaces umlaute
    >>> original_username = "Mädchën José"
    >>> cleaned_username = replace_umlaute_and_remove_accents(original_username)
    >>> print(cleaned_username)
    Maedchen Jose
    >>> cleaned_username = replace_umlaute_and_remove_accents("Bernt Straßenbahner")
    >>> print(cleaned_username)
    Bernt Strassenbahner
    """
    username = username.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')


    norm_txt = unicodedata.normalize('NFD', username)
    shaved = ''.join(c for c in norm_txt if not unicodedata.combining(c))
    username = unicodedata.normalize('NFC', shaved)

    return username


def get_classes_from_class_file(path:str) -> Generator[str, None, None]:
    """
    a generator that reads from excel file
    :param path: the path of excel file
    :return: tuple(class_name, room, kv)
    """
    try:
        wb = load_workbook(Path(path), read_only=True)
        ws = wb[wb.sheetnames[0]]
        logger.debug("Reading class excel file")
        for row in ws.iter_rows(min_row=2):
            class_name = row[0].value
            room = row[1].value
            kv = row[2].value
            if class_name is None:
                break
            yield class_name, room, kv
    except FileNotFoundError:
        logger.critical(f"Couldn't find file {path}")
    except Exception:
        logger.critical(f"Couldn't access file {path}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument("-q", "--quiet", help='Only Log Errors', required=False, action='store_true')
    parser.add_argument("-v", "--verbose", help='Log Verbose Debug', required=False, action='store_true')


    args = parser.parse_args()
    global logger
    logger = logging.getLogger('my_logger')
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    elif args.quiet:
        logger.setLevel(logging.ERROR)
    else:
        logger.setLevel(logging.INFO)

    file_handler = RotatingFileHandler('user_logfile.log', maxBytes=10000, backupCount=5)
    file_handler.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s %(message)s'))
    stream_handler = StreamHandler()
    stream_formatter = logging.Formatter('%(levelname)s - %(message)s')
    stream_handler.setFormatter(stream_formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.info("Logging turned on " + str(logger.level))
    # create_user_from_name_file("../../res/Namen.xlsx")
    create_class_user_files("../../res/Klassenraeume_2023.xlsx")

