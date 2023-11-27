import argparse
import random
import string
from pathlib import Path
from typing import Generator

import openpyxl
import unicodedata
from openpyxl import load_workbook
import logging
from logging.handlers import RotatingFileHandler
from logging import StreamHandler

def create_user_from_name_file(path:str):
    """
    creates a user class bash file, a delteion file and a userlist
    :param path: The excel file
    """
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Name'
        sheet['B1'] = 'Password'
        with open("create_real_users.sh", 'w') as file, open("delete_real_users.sh", 'w') as file_delete, \
                open("userlist.txt", 'w') as user_list_file:
            file.write("#! /bin/sh\n")
            file_delete.write("#! /bin/sh\n")
            file.write("groupadd schueler\n")
            user_dict = dict()
            index = 2
            for first_name, last_name, group, school_class in get_people_from_name_file(path):
                username = replace_umlaute_and_remove_accents(last_name).replace(" ", "_")
                if username in user_dict:
                    user_dict[username] += 1
                    username = username+ f"_{user_dict[username]}"
                else:
                    user_dict[username] = 0
                command = f"useradd -d \"/home/klassen/{username}\" -g schueler -c \"{username} - {school_class}\" " \
                          f"-s /bin/bash -G cdrom,plugdev,sambashare {username}\n"
                password = get_random_password()
                file.write(command)
                logger.debug(f"User created: {username}")
                passwd_command = f'echo {username}:{password} | chpasswd\n'
                file.write(passwd_command)
                logger.debug(f"User password created for {username}")
                file_delete.write(f'userdel -r {username}\n')
                logger.debug(f"User deletion written for {username}")
                user_list_file.write(f'username: {username}, password: {password}\n')
                logger.debug(f"User \"{username}\" added to list")
                sheet[f'A{index}'] = username
                sheet[f'B{index}'] = password
                logger.debug(f"Wrote user {username} in excel")
                index+=1
        workbook.save('users.xlsx')
    except Exception:
        logger.critical("Couldn't write in one of the Files")

def get_people_from_name_file(path) -> Generator[str, None, None]:
    """
        a generator that reads from excel file
        :param path: the path of excel file
        :return: tuple(first_name, last_name, group, school_class)
        """
    try:
        wb = load_workbook(Path(path), read_only=True)
        ws = wb[wb.sheetnames[0]]
        logger.debug("Read from excel")
        for row in ws.iter_rows(min_row=2):
            first_name = row[0].value
            last_name = row[1].value
            group = row[2].value
            school_class = row[3].value
            yield first_name, last_name, group, school_class
    except Exception:
        logger.critical("Couln't read excel file")


def get_random_password() -> str:
    """
    :return: 12 letter password
    """
    return "".join([random.choice(string.ascii_letters + '!%&(),._-=^#' + string.digits) if i > 0 else random.choice(string.ascii_letters) for i in range(12)])


def replace_umlaute_and_remove_accents(username):
    """
    >>> original_username = "Mädchën José"
    >>> cleaned_username = replace_umlaute_and_remove_accents(original_username)
    >>> print(cleaned_username)
    Maedchen Jose
    >>> cleaned_username = replace_umlaute_and_remove_accents("Bernt Straßenbahner")
    >>> print(cleaned_username)
    Bernt Strassenbahner
    """
    username = username.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss').replace(r"\s*", "_")


    norm_txt = unicodedata.normalize('NFD', username)
    shaved = ''.join(c for c in norm_txt if not unicodedata.combining(c))
    username = unicodedata.normalize('NFC', shaved)

    return username

if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument("-q", "--quiet", help='Only Log Errors', required=False, action='store_true')
    parser.add_argument("-v", "--verbose", help='Log Verbose Debug', required=False, action='store_true')


    args = parser.parse_args()
    global logger
    logger = logging.getLogger('my_logger')
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    elif args.quiet:
        logger.setLevel(logging.ERROR)
    else:
        logger.setLevel(logging.INFO)

    file_handler = RotatingFileHandler('user_logfile.log', maxBytes=10000, backupCount=5)
    file_handler.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s %(message)s'))
    stream_handler = StreamHandler()
    stream_formatter = logging.Formatter('%(levelname)s - %(message)s')
    stream_handler.setFormatter(stream_formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.info("Logging turned on " + str(logger.level))
    # create_user_from_name_file("../../res/Namen.xlsx")
    create_user_from_name_file("../../res/Namen.xlsx")